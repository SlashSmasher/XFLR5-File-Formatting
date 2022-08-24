from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

number_worksheets = int(input("Insert the number of calculations: "))
wb = Workbook()
sheet_index = 0

number = number_worksheets

while number > 1:
    ws = wb.create_sheet()
    number -= 1

sheet_names = wb.sheetnames

while number_worksheets > 0:
    active_sheet = sheet_names[sheet_index]
    ws = wb[active_sheet]
    row = []
    iter_row = 2
    file_directory = input("Input file directory: ")

    with open(file_directory, "r") as f:
        contents = f.readlines()

    airfoil = contents[2]
    airfoil = airfoil.split(":")
    airfoil = airfoil[1]
    airfoil = airfoil[1:]
    airfoil = airfoil.replace("\n", "")
    ws.title = airfoil


    for times in range(0, 9):
        contents.pop(0)

    for times in range(0, 2):
        number_of_items = len(contents)
        contents.pop(number_of_items - 1)


    headings = contents[0].split(",")
    ws.append(headings)

    reformat_numbers = contents[1:]
    for item in reformat_numbers:
        item = item.split(",")
        for item in item:
            item = item.replace(".", ",")
            has_spaces = item.count(" ")
            if has_spaces != 0:
                item = item.replace(" ", "", has_spaces)
            row.append(item)
            row_elements = len(row)
            if row_elements == 12:
                ws.append(row)
                row.clear()

    while True:
        cell = "A" + str(iter_row)
        if ws[cell].value == None:
            break
        else:
            iter_row += 1


    for col in range(1, 12):
        for row in range(2, iter_row):
            lcol = get_column_letter(col)
            cell = lcol + str(row)
            value = ws[cell].value
            value = value.replace(",", ".")
            ws[cell].value = float(value)

    with open(file_directory, "r") as f:
        analisys = f.readlines()

    info = analisys[7]
    info = info.replace("     ", ",")
    index = info.index(",")
    info = info[:index] + ";" + info[index + 1:]
    index = info.index(",")
    info = info[:index] + " " + info[index + 1:]
    index = info.index(",")
    info = info[:index] + ";" + info[index + 1:]
    info = info[1:]
    info = info.split(";")
    mach = info[0]
    re = info[1]
    ncrit = info[2]
    ncrit = ncrit[:-1]
    mach = mach.replace("   ", " ")
    ncrit = ncrit.replace("   ", " ")
    re = re.replace("Re", "R")
    iexp = re.index("e")
    exp = re[iexp + 2:]
    exp = 10 ** int(exp)
    nre = re[:iexp - 1]
    nindex = nre.index("0")
    nre = nre[nindex:]
    reynolds = float(nre) * int(exp)
    reynolds = str(int(reynolds))
    eqi = re.index("=")
    substitute = re[eqi + 2:]
    re = re.replace(substitute, reynolds)
    re = re.replace("R", "Reynolds")

    ws["N2"].value = mach
    ws["N3"].value = re
    ws["N4"].value = ncrit
    ws.column_dimensions['N'].width = 20

    number_worksheets -= 1
    sheet_index += 1


save_file_directory = input("Input the directory where you want to save the file: ")
save_file_name = input("Choose the name you want to give to your file: ")
save_file = save_file_directory + "/" + save_file_name + ".xlsx"
wb.save(save_file)




